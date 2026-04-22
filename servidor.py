# ============================================================
# SERVIDOR STRATIVA — Versão Final
# ============================================================
# COMO USAR:
#   1. python -m pip install flask flask-socketio openpyxl
#   2. python servidor.py
#   3. Formulário → http://localhost:5000
#   4. Admin      → http://localhost:5000/admin
#        utilizador: ADMIM
#        password:   GDNM
# ============================================================

from flask import Flask, request, jsonify, send_file, render_template_string, redirect, session
from flask_socketio import SocketIO
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from functools import wraps
import os
import eventlet
eventlet.monkey_patch()

app = Flask(__name__)
app.secret_key = "strativa_chave_secreta_2026"

# SocketIO permite enviar dados em tempo real para o browser
socketio = SocketIO(app, cors_allowed_origins="*")

ADMIN_UTILIZADOR = "ADMIN"
ADMIN_PASSWORD   = "GDNM"
NOME_BASE_DADOS  = "strativa.db"


# ------------------------------------------------------------
# BASE DE DADOS
# ------------------------------------------------------------
def criar_tabela():
    con = sqlite3.connect(NOME_BASE_DADOS)
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS respostas (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            data_envio    TEXT,
            nome_negocio  TEXT,
            setor         TEXT,
            tempo_mercado TEXT,
            faturamento   TEXT,
            digital       TEXT,
            estrutura     TEXT,
            fluxo         TEXT,
            objetivos     TEXT,
            visao         TEXT,
            investimento  TEXT,
            ambicao       TEXT
        )
    """)
    con.commit()
    con.close()

def buscar_todas_respostas():
    con = sqlite3.connect(NOME_BASE_DADOS)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute("SELECT * FROM respostas ORDER BY id DESC")
    rows = [dict(r) for r in cur.fetchall()]
    con.close()
    return rows

def contar_respostas():
    con = sqlite3.connect(NOME_BASE_DADOS)
    cur = con.cursor()
    cur.execute("SELECT COUNT(*) FROM respostas")
    total = cur.fetchone()[0]
    con.close()
    return total


# ------------------------------------------------------------
# PROTEÇÃO DO ADMIN
# ------------------------------------------------------------
def requer_login(f):
    @wraps(f)
    def verificar(*args, **kwargs):
        if not session.get("admin_logado"):
            return redirect("/admin/login")
        return f(*args, **kwargs)
    return verificar


# ============================================================
# ROTAS PÚBLICAS
# ============================================================

# O formulário está embutido aqui para não depender de ficheiros externos
@app.route("/")
def pagina_inicial():
    return render_template_string(FORMULARIO_HTML)


@app.route("/enviar", methods=["POST"])
def receber_formulario():
    dados = request.get_json()
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    con = sqlite3.connect(NOME_BASE_DADOS)
    cur = con.cursor()
    cur.execute("""
        INSERT INTO respostas (
            data_envio, nome_negocio, setor, tempo_mercado,
            faturamento, digital, estrutura, fluxo,
            objetivos, visao, investimento, ambicao
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        agora,
        dados.get("nome_negocio", ""),
        dados.get("setor", ""),
        dados.get("tempo_mercado", ""),
        dados.get("faturamento", ""),
        dados.get("digital", ""),
        dados.get("estrutura", ""),
        dados.get("fluxo", ""),
        dados.get("objetivos", ""),
        dados.get("visao", ""),
        dados.get("investimento", ""),
        dados.get("ambicao", "")
    ))
    con.commit()
    novo_id = cur.lastrowid
    con.close()

    # Busca a resposta recém inserida
    con = sqlite3.connect(NOME_BASE_DADOS)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute("SELECT * FROM respostas WHERE id = ?", (novo_id,))
    nova = dict(cur.fetchone())
    con.close()

    # ⚡ Envia em tempo real para todos os admins abertos no browser
    socketio.emit("nova_resposta", {
        "resposta": nova,
        "total": contar_respostas()
    })

    return jsonify({"sucesso": True})


# ============================================================
# ROTAS DE ADMINISTRAÇÃO
# ============================================================

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    erro = ""
    if request.method == "POST":
        utilizador = request.form.get("utilizador", "")
        password   = request.form.get("password", "")
        if utilizador == ADMIN_UTILIZADOR and password == ADMIN_PASSWORD:
            session["admin_logado"] = True
            return redirect("/admin")
        else:
            erro = "Utilizador ou password incorretos."
    return render_template_string(LOGIN_HTML, erro=erro)


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect("/admin/login")


@app.route("/admin")
@requer_login
def admin_painel():
    respostas = buscar_todas_respostas()
    total     = contar_respostas()
    return render_template_string(ADMIN_HTML, respostas=respostas, total=total)


@app.route("/admin/exportar")
@requer_login
def admin_exportar():
    respostas = buscar_todas_respostas()

    colunas = [
        "ID", "Data de Envio", "Nome do Negócio", "Setor",
        "Tempo de Mercado", "Faturamento", "Presença Digital",
        "Estrutura", "Fluxo de Clientes", "Objetivos",
        "Visão Estratégica", "Investimento em Marketing", "Ambição"
    ]
    chaves = [
        "id", "data_envio", "nome_negocio", "setor",
        "tempo_mercado", "faturamento", "digital",
        "estrutura", "fluxo", "objetivos",
        "visao", "investimento", "ambicao"
    ]

    livro = openpyxl.Workbook()
    folha = livro.active
    folha.title = "Respostas Strativa"

    folha.append(colunas)
    for celula in folha[1]:
        celula.font      = Font(bold=True, color="FFFFFF")
        celula.fill      = PatternFill("solid", fgColor="1a1a1a")
        celula.alignment = Alignment(horizontal="center")

    for r in respostas:
        folha.append([r.get(k, "") for k in chaves])

    for col in folha.columns:
        largura = max(len(str(c.value or "")) for c in col) + 4
        folha.column_dimensions[col[0].column_letter].width = min(largura, 50)

    nome = "respostas_strativa.xlsx"
    livro.save(nome)
    return send_file(nome, as_attachment=True)


@app.route("/admin/apagar/<int:id>", methods=["POST"])
@requer_login
def admin_apagar(id):
    con = sqlite3.connect(NOME_BASE_DADOS)
    cur = con.cursor()
    cur.execute("DELETE FROM respostas WHERE id = ?", (id,))
    con.commit()
    con.close()
    # Notifica o admin em tempo real da remoção
    socketio.emit("resposta_apagada", {
        "id": id,
        "total": contar_respostas()
    })
    return redirect("/admin")


# ============================================================
# HTML DO FORMULÁRIO PÚBLICO
# ============================================================
FORMULARIO_HTML = """<!DOCTYPE html>
<html class="dark" lang="pt-BR">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Strativa — Avaliação Estratégica</title>
  <script src="https://cdn.tailwindcss.com?plugins=forms,container-queries"></script>
  <link href="https://fonts.googleapis.com/css2?family=Manrope:wght@200;300;400;500;600;700;800&display=swap" rel="stylesheet"/>
  <link href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:wght,FILL@100..700,0..1&display=swap" rel="stylesheet"/>
  <script>
    tailwind.config = {
      darkMode: "class",
      theme: { extend: { colors: {
        "surface-container-highest":"#252626","background":"#0e0e0e",
        "surface-container":"#191a1a","outline-variant":"#484848",
        "surface":"#0e0e0e","on-background":"#e7e5e4",
        "on-surface-variant":"#acabaa","surface-container-low":"#131313",
        "surface-container-high":"#1f2020","on-surface":"#e7e5e4",
        "secondary-container":"#3a3c3c","on-primary":"#3f4041",
        "secondary":"#9e9e9e","primary":"#c6c6c7","outline":"#767575"
      }, fontFamily: { headline:["Manrope"], body:["Manrope"] } } }
    }
  </script>
  <style>
    body { font-family:'Manrope',sans-serif; background:#0e0e0e; color:#e7e5e4; min-height:max(884px,100dvh); }
    .material-symbols-outlined { font-variation-settings:'FILL' 0,'wght' 400,'GRAD' 0,'opsz' 24; }
    .custom-radio:checked + label { background-color:#252626; border-color:#c6c6c7; color:#fff; }
    .metallic-gradient { background:linear-gradient(135deg,#c6c6c7 0%,#454747 100%); }
    input:focus { outline:none!important; }
    #popup-overlay { display:none; position:fixed; inset:0; background:rgba(0,0,0,.75); z-index:100; align-items:center; justify-content:center; }
    #popup-overlay.visivel { display:flex; }
    #popup-erro { display:none; position:fixed; inset:0; background:rgba(0,0,0,.75); z-index:100; align-items:center; justify-content:center; }
    #popup-erro.visivel { display:flex; }
  </style>
</head>
<body class="bg-background text-on-surface antialiased">

<!-- POPUP SUCESSO -->
<div id="popup-overlay">
  <div class="bg-[#191a1a] border border-[#484848]/30 rounded-xl p-10 max-w-sm mx-6 text-center space-y-4">
    <div class="text-4xl">✅</div>
    <p class="text-lg font-bold text-primary">Avaliação Enviada!</p>
    <p class="text-secondary text-sm">Os seus dados foram registados com sucesso.</p>
    <button onclick="fecharPopup()" class="mt-4 px-8 py-3 rounded border border-primary/40 text-primary text-sm uppercase tracking-widest hover:bg-primary/10 transition-colors">Fechar</button>
  </div>
</div>

<!-- POPUP ERRO -->
<div id="popup-erro">
  <div class="bg-[#191a1a] border border-red-900/40 rounded-xl p-10 max-w-sm mx-6 text-center space-y-4">
    <div class="text-4xl">❌</div>
    <p class="text-lg font-bold text-red-400">Erro ao Enviar</p>
    <p class="text-secondary text-sm">Certifica-te que o servidor está a correr.<br/><code class="text-primary">python servidor.py</code></p>
    <button onclick="document.getElementById('popup-erro').classList.remove('visivel')" class="mt-4 px-8 py-3 rounded border border-red-900/40 text-red-400 text-sm uppercase tracking-widest">Fechar</button>
  </div>
</div>

<!-- CABEÇALHO -->
<header class="bg-gradient-to-b from-[#0e0e0e] to-transparent sticky top-0 z-50">
  <div class="flex justify-between items-center w-full px-6 py-6 max-w-screen-2xl mx-auto">
    <div class="text-2xl font-bold tracking-tight text-zinc-300 uppercase">Strativa</div>
  </div>
</header>

<main class="max-w-2xl mx-auto px-6 py-12 space-y-24 pb-32">

  <section class="space-y-4">
    <h1 class="text-5xl font-extrabold tracking-tighter text-on-surface leading-tight">
      Formulário de <span class="text-primary italic">Avaliação</span> Strativa
    </h1>
    <p class="text-secondary max-w-md text-lg leading-relaxed">
      Um diagnóstico estratégico para elevar o posicionamento e a performance do seu negócio no mercado.
    </p>
  </section>

  <form id="formulario-strativa" class="space-y-20">

    <!-- 01 Informações Básicas -->
    <div class="space-y-8">
      <div class="space-y-2">
        <span class="text-xs font-bold tracking-[0.2em] text-primary uppercase">01. Informações Básicas</span>
        <div class="h-px w-12 bg-primary"></div>
      </div>
      <div class="grid gap-8">
        <div>
          <label class="block text-[10px] uppercase tracking-widest text-secondary mb-1">Nome do Negócio</label>
          <input name="nome_negocio" type="text" placeholder="Ex: Strativa Consulting" class="w-full bg-transparent border-b border-outline-variant/40 py-3 text-on-surface focus:border-primary transition-colors placeholder:text-zinc-800"/>
        </div>
        <div>
          <label class="block text-[10px] uppercase tracking-widest text-secondary mb-1">Setor de Atuação</label>
          <input name="setor" type="text" placeholder="Ex: Consultoria, Varejo, Saúde" class="w-full bg-transparent border-b border-outline-variant/40 py-3 text-on-surface focus:border-primary transition-colors placeholder:text-zinc-800"/>
        </div>
        <div>
          <label class="block text-[10px] uppercase tracking-widest text-secondary mb-1">Tempo de Mercado</label>
          <input name="tempo_mercado" type="text" placeholder="Ex: 5 anos" class="w-full bg-transparent border-b border-outline-variant/40 py-3 text-on-surface focus:border-primary transition-colors placeholder:text-zinc-800"/>
        </div>
      </div>
    </div>

    <!-- 02 Faturamento -->
    <div class="space-y-8">
      <div class="space-y-2">
        <span class="text-xs font-bold tracking-[0.2em] text-primary uppercase">02. Faturamento</span>
        <div class="h-px w-12 bg-primary"></div>
      </div>
      <div class="grid gap-3">
        <div><input class="hidden custom-radio" type="radio" id="fat1" name="faturamento" value="Até 500.000 Kz / mês"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="fat1">Até 500.000 Kz / mês</label></div>
        <div><input class="hidden custom-radio" type="radio" id="fat2" name="faturamento" value="De 500.000 Kz a 2.000.000 Kz / mês"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="fat2">De 500.000 Kz a 2.000.000 Kz / mês</label></div>
        <div><input class="hidden custom-radio" type="radio" id="fat3" name="faturamento" value="Acima de 2.000.000 Kz / mês"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="fat3">Acima de 2.000.000 Kz / mês</label></div>
      </div>
    </div>

    <!-- 03 Presença Digital -->
    <div class="space-y-8">
      <div class="space-y-2">
        <span class="text-xs font-bold tracking-[0.2em] text-primary uppercase">03. Presença Digital</span>
        <div class="h-px w-12 bg-primary"></div>
      </div>
      <div class="grid gap-3">
        <div><input class="hidden custom-radio" type="radio" id="dig1" name="digital" value="Inexistente ou muito irregular"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="dig1">Inexistente ou muito irregular (postagens esporádicas)</label></div>
        <div><input class="hidden custom-radio" type="radio" id="dig2" name="digital" value="Frequente mas com baixo engajamento"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="dig2">Frequente, mas com baixo engajamento ou qualidade visual amadora</label></div>
        <div><input class="hidden custom-radio" type="radio" id="dig3" name="digital" value="Profissional e estratégica"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="dig3">Profissional, estratégica e com crescimento constante</label></div>
      </div>
    </div>

    <!-- 04 Estrutura -->
    <div class="space-y-8">
      <div class="space-y-2">
        <span class="text-xs font-bold tracking-[0.2em] text-primary uppercase">04. Estrutura Organizacional</span>
        <div class="h-px w-12 bg-primary"></div>
      </div>
      <div class="grid gap-3">
        <div><input class="hidden custom-radio" type="radio" id="est1" name="estrutura" value="Eu faço tudo (Eupresa)"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="est1">Eu faço tudo (Eupresa)</label></div>
        <div><input class="hidden custom-radio" type="radio" id="est2" name="estrutura" value="Equipa pequena sem processos definidos"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="est2">Equipa pequena, mas sem processos definidos</label></div>
        <div><input class="hidden custom-radio" type="radio" id="est3" name="estrutura" value="Estrutura sólida com departamentos e processos claros"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="est3">Estrutura sólida com departamentos e processos claros</label></div>
      </div>
    </div>

    <!-- 05 Fluxo -->
    <div class="space-y-8">
      <div class="space-y-2">
        <span class="text-xs font-bold tracking-[0.2em] text-primary uppercase">05. Fluxo de Clientes</span>
        <div class="h-px w-12 bg-primary"></div>
      </div>
      <div class="grid gap-3">
        <div><input class="hidden custom-radio" type="radio" id="flu1" name="fluxo" value="Dependo de indicações e sorte"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="flu1">Dependo exclusivamente de indicações e "sorte"</label></div>
        <div><input class="hidden custom-radio" type="radio" id="flu2" name="fluxo" value="Tenho movimento mas instável"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="flu2">Tenho movimento, mas é instável e imprevisível</label></div>
        <div><input class="hidden custom-radio" type="radio" id="flu3" name="fluxo" value="Sistema de vendas recorrente"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="flu3">Possuo um sistema de vendas que atrai clientes recorrentemente</label></div>
      </div>
    </div>

    <!-- 06 Objetivos -->
    <div class="space-y-8">
      <div class="space-y-2">
        <span class="text-xs font-bold tracking-[0.2em] text-primary uppercase">06. Objetivos & Crescimento</span>
        <div class="h-px w-12 bg-primary"></div>
      </div>
      <div>
        <label class="block text-[10px] uppercase tracking-widest text-secondary mb-3">Qual o seu principal objetivo para os próximos 6 meses?</label>
        <textarea name="objetivos" rows="4" placeholder="Descreva suas metas de faturamento, expansão ou posicionamento..." class="w-full bg-surface-container-low border border-outline-variant/15 p-4 rounded text-on-surface focus:border-primary transition-colors placeholder:text-zinc-800 resize-none"></textarea>
      </div>
    </div>

    <!-- 07 Visão -->
    <div class="space-y-8">
      <div class="space-y-2">
        <span class="text-xs font-bold tracking-[0.2em] text-primary uppercase">07. Visão Estratégica</span>
        <div class="h-px w-12 bg-primary"></div>
      </div>
      <div class="grid gap-3">
        <div><input class="hidden custom-radio" type="radio" id="vis1" name="visao" value="Vejo marketing como custo"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="vis1">Vejo Marketing como um custo que tento evitar</label></div>
        <div><input class="hidden custom-radio" type="radio" id="vis2" name="visao" value="Necessário mas com medo de arriscar"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="vis2">Entendo que é necessário, mas tenho medo de arriscar alto</label></div>
        <div><input class="hidden custom-radio" type="radio" id="vis3" name="visao" value="Investimento essencial para escala"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="vis3">Encaro como investimento essencial para a escala do negócio</label></div>
      </div>
    </div>

    <!-- 08 Investimento -->
    <div class="space-y-8">
      <div class="space-y-2">
        <span class="text-xs font-bold tracking-[0.2em] text-primary uppercase">08. Investimento em Marketing</span>
        <div class="h-px w-12 bg-primary"></div>
      </div>
      <div class="grid gap-3">
        <div><input class="hidden custom-radio" type="radio" id="inv1" name="investimento" value="Menos de 50.000 Kz / mês"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="inv1">Nenhum ou menos de 50.000 Kz / mês</label></div>
        <div><input class="hidden custom-radio" type="radio" id="inv2" name="investimento" value="De 50.000 a 200.000 Kz / mês"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="inv2">De 50.000 Kz a 200.000 Kz / mês</label></div>
        <div><input class="hidden custom-radio" type="radio" id="inv3" name="investimento" value="Acima de 200.000 Kz / mês"/><label class="block p-5 rounded border border-outline-variant/15 bg-surface-container-low hover:bg-surface-container-high transition-all cursor-pointer text-sm" for="inv3">Acima de 200.000 Kz / mês</label></div>
      </div>
    </div>

    <!-- 09 Ambição -->
    <div class="space-y-8">
      <div class="space-y-2">
        <span class="text-xs font-bold tracking-[0.2em] text-primary uppercase">09. Ambição Profissional</span>
        <div class="h-px w-12 bg-primary"></div>
      </div>
      <div>
        <label class="block text-[10px] uppercase tracking-widest text-secondary mb-3">Onde você se vê como líder e empresário em 2 anos?</label>
        <textarea name="ambicao" rows="5" placeholder="Descreva seu nível de compromisso e a visão de futuro para sua marca pessoal e empresa..." class="w-full bg-surface-container-low border border-outline-variant/15 p-4 rounded text-on-surface focus:border-primary transition-colors placeholder:text-zinc-800 resize-none"></textarea>
      </div>
    </div>

    <!-- BOTÃO ENVIAR -->
    <button id="botao-enviar" type="submit" class="w-full metallic-gradient py-5 rounded font-bold text-on-primary tracking-widest uppercase text-sm shadow-[0_10px_30px_rgba(198,198,199,0.1)] active:scale-[0.98] transition-transform">
      Enviar Avaliação Estratégica
    </button>

  </form>

  <!-- RODAPÉ -->
  <footer class="pt-24 pb-12 border-t border-outline-variant/10">
    <div class="space-y-8">
      <div class="space-y-1">
        <p class="text-2xl font-light tracking-tight text-primary">STRATIVA</p>
        <p class="text-[10px] uppercase tracking-[0.3em] text-secondary">Growth & Strategy</p>
      </div>
      <div class="flex items-start gap-12 text-[10px] uppercase tracking-widest text-secondary">
        <div class="space-y-2"><span class="block text-primary/40">Localização</span><span>Luanda, Angola</span></div>
        <div class="space-y-2"><span class="block text-primary/40">Consultoria</span><span>Premium Growth</span></div>
      </div>
      <div class="pt-12 text-center">
        <span class="text-[8px] tracking-[0.5em] text-outline opacity-50 uppercase">Strativa © 2026 • All Rights Reserved</span>
      </div>
    </div>
  </footer>

</main>

<script>
  const formulario = document.getElementById("formulario-strativa");
  const botao      = document.getElementById("botao-enviar");

  formulario.addEventListener("submit", async function(e) {
    e.preventDefault();
    botao.textContent = "A enviar...";
    botao.disabled = true;

    const dados = {
      nome_negocio:  formulario.nome_negocio.value,
      setor:         formulario.setor.value,
      tempo_mercado: formulario.tempo_mercado.value,
      faturamento:   formulario.faturamento ? formulario.faturamento.value : "",
      digital:       formulario.digital     ? formulario.digital.value     : "",
      estrutura:     formulario.estrutura   ? formulario.estrutura.value   : "",
      fluxo:         formulario.fluxo       ? formulario.fluxo.value       : "",
      objetivos:     formulario.objetivos.value,
      visao:         formulario.visao       ? formulario.visao.value       : "",
      investimento:  formulario.investimento? formulario.investimento.value: "",
      ambicao:       formulario.ambicao.value
    };

    try {
      const resposta = await fetch("/enviar", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(dados)
      });
      const resultado = await resposta.json();
      if (resultado.sucesso) {
        document.getElementById("popup-overlay").classList.add("visivel");
        formulario.reset();
      } else {
        document.getElementById("popup-erro").classList.add("visivel");
      }
    } catch(err) {
      document.getElementById("popup-erro").classList.add("visivel");
    }

    botao.textContent = "Enviar Avaliação Estratégica";
    botao.disabled = false;
  });

  function fecharPopup() {
    document.getElementById("popup-overlay").classList.remove("visivel");
  }
</script>
</body>
</html>"""


# ============================================================
# HTML DO LOGIN
# ============================================================
LOGIN_HTML = """<!DOCTYPE html>
<html lang="pt">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Strativa — Admin</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <link href="https://fonts.googleapis.com/css2?family=Manrope:wght@300;400;600;700;800&display=swap" rel="stylesheet"/>
  <style>body { font-family:'Manrope',sans-serif; background:#0a0a0a; color:#e7e5e4; }</style>
</head>
<body class="min-h-screen flex items-center justify-center px-6">
  <div class="w-full max-w-sm space-y-10">
    <div class="text-center space-y-2">
      <p class="text-2xl font-bold tracking-[0.3em] text-zinc-300 uppercase">Strativa</p>
      <p class="text-[10px] tracking-[0.2em] text-zinc-600 uppercase">Painel de Administração</p>
    </div>
    <form method="POST" class="space-y-6">
      {% if erro %}
      <div class="bg-red-950/40 border border-red-900/40 rounded p-4 text-red-400 text-sm text-center">{{ erro }}</div>
      {% endif %}
      <div class="space-y-1">
        <label class="block text-[10px] uppercase tracking-widest text-zinc-500">Utilizador</label>
        <input name="utilizador" type="text" autocomplete="username" class="w-full bg-transparent border-b border-zinc-800 py-3 text-zinc-200 focus:border-zinc-400 outline-none transition-colors" placeholder="nome de utilizador"/>
      </div>
      <div class="space-y-1">
        <label class="block text-[10px] uppercase tracking-widest text-zinc-500">Password</label>
        <input name="password" type="password" autocomplete="current-password" class="w-full bg-transparent border-b border-zinc-800 py-3 text-zinc-200 focus:border-zinc-400 outline-none transition-colors" placeholder="••••••••"/>
      </div>
      <button type="submit" class="w-full py-4 bg-zinc-200 text-zinc-900 font-bold uppercase tracking-widest text-sm rounded hover:bg-white transition-colors">Entrar</button>
    </form>
    <p class="text-center text-[9px] tracking-widest text-zinc-700 uppercase">Acesso restrito · Não partilhes este link</p>
  </div>
</body>
</>"""


# ============================================================
# HTML DO PAINEL ADMIN (com tempo real via SocketIO)
# ============================================================
ADMIN_HTML = """<!DOCTYPE html>
<html lang="pt">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Strativa — Admin</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <script src="https://cdn.socket.io/4.7.5/socket.io.min.js"></script>
  <link href="https://fonts.googleapis.com/css2?family=Manrope:wght@300;400;600;700;800&display=swap" rel="stylesheet"/>
  <link href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:wght,FILL@100..700,0..1&display=swap" rel="stylesheet"/>
  <style>
    body { font-family:'Manrope',sans-serif; background:#0a0a0a; color:#e7e5e4; }
    .material-symbols-outlined { font-variation-settings:'FILL' 0,'wght' 400,'GRAD' 0,'opsz' 24; vertical-align:middle; }
    ::-webkit-scrollbar { width:4px; height:4px; }
    ::-webkit-scrollbar-track { background:#111; }
    ::-webkit-scrollbar-thumb { background:#333; border-radius:4px; }
    #modal { display:none; position:fixed; inset:0; background:rgba(0,0,0,.85); z-index:50; align-items:center; justify-content:center; padding:1rem; }
    #modal.aberto { display:flex; }
    /* Animação da linha nova */
    @keyframes entrar { from { opacity:0; background:#1a2a1a; } to { opacity:1; background:transparent; } }
    .linha-nova { animation: entrar 1.5s ease forwards; }
    /* Badge de novo */
    @keyframes pulsar { 0%,100%{opacity:1} 50%{opacity:.4} }
    .badge-novo { animation: pulsar 1.5s ease 3; }
    /* Notificação toast */
    #toast { position:fixed; bottom:2rem; right:2rem; z-index:99; transform:translateY(100px); opacity:0; transition:all .4s ease; }
    #toast.visivel { transform:translateY(0); opacity:1; }
  </style>
</head>
<body class="min-h-screen">

<!-- TOAST DE NOTIFICAÇÃO -->
<div id="toast">
  <div class="bg-green-900/80 border border-green-700/50 text-green-300 px-6 py-4 rounded-xl text-sm font-medium flex items-center gap-3 backdrop-blur">
    <span class="material-symbols-outlined text-green-400">notifications</span>
    <span id="toast-msg">Nova resposta recebida!</span>
  </div>
</div>

<!-- MODAL DETALHE -->
<div id="modal">
  <div class="bg-[#131313] border border-zinc-800/50 rounded-xl w-full max-w-2xl max-h-[90vh] overflow-y-auto p-8 space-y-6 relative">
    <button onclick="fecharModal()" class="absolute top-4 right-4 text-zinc-500 hover:text-white transition-colors">
      <span class="material-symbols-outlined">close</span>
    </button>
    <p class="text-xs uppercase tracking-widest text-zinc-500">Detalhe da Resposta</p>
    <div id="modal-conteudo" class="space-y-4"></div>
  </div>
</div>

<!-- CABEÇALHO -->
<header class="border-b border-zinc-900 px-8 py-5 flex items-center justify-between sticky top-0 bg-[#0a0a0a]/95 backdrop-blur z-40">
  <div class="flex items-center gap-4">
    <p class="text-lg font-bold tracking-[0.2em] text-zinc-300 uppercase">Strativa</p>
    <span class="text-[10px] bg-zinc-800 text-zinc-400 px-2 py-1 rounded tracking-widest uppercase">Admin</span>
    <!-- Indicador de ligação em tempo real -->
    <div class="flex items-center gap-1.5">
      <div id="dot-ligado" class="w-2 h-2 rounded-full bg-green-500"></div>
      <span id="txt-ligado" class="text-[10px] text-zinc-500 uppercase tracking-widest">Em tempo real</span>
    </div>
  </div>
  <div class="flex items-center gap-3">
    <a href="/admin/exportar" class="flex items-center gap-2 px-4 py-2 bg-zinc-200 text-zinc-900 text-xs font-bold uppercase tracking-widest rounded hover:bg-white transition-colors">
      <span class="material-symbols-outlined text-sm">download</span>Exportar Excel
    </a>
    <a href="/admin/logout" class="flex items-center gap-2 px-4 py-2 border border-zinc-800 text-zinc-500 text-xs uppercase tracking-widest rounded hover:border-zinc-600 hover:text-zinc-300 transition-colors">
      <span class="material-symbols-outlined text-sm">logout</span>Sair
    </a>
  </div>
</header>

<main class="max-w-7xl mx-auto px-8 py-10 space-y-8">

  <!-- Estatísticas -->
  <div class="grid grid-cols-1 sm:grid-cols-3 gap-4">
    <div class="bg-[#111] border border-zinc-800/40 rounded-xl p-6 space-y-1">
      <p class="text-[10px] uppercase tracking-widest text-zinc-500">Total de Respostas</p>
      <p id="contador-total" class="text-4xl font-bold text-zinc-200">{{ total }}</p>
    </div>
    <div class="bg-[#111] border border-zinc-800/40 rounded-xl p-6 space-y-1">
      <p class="text-[10px] uppercase tracking-widest text-zinc-500">Última Entrada</p>
      <p id="ultima-data" class="text-sm font-semibold text-zinc-300">
        {% if respostas %}{{ respostas[0]['data_envio'] }}{% else %}—{% endif %}
      </p>
    </div>
    <div class="bg-[#111] border border-zinc-800/40 rounded-xl p-6 space-y-1">
      <p class="text-[10px] uppercase tracking-widest text-zinc-500">Exportar Dados</p>
      <a href="/admin/exportar" class="text-sm font-semibold text-zinc-300 hover:text-white underline underline-offset-4 transition-colors">Descarregar .xlsx →</a>
    </div>
  </div>

  <!-- Pesquisa -->
  <div class="relative">
    <span class="material-symbols-outlined absolute left-4 top-1/2 -translate-y-1/2 text-zinc-600">search</span>
    <input id="pesquisa" type="text" placeholder="Pesquisar por nome, setor..." oninput="filtrar()"
      class="w-full bg-[#111] border border-zinc-800/40 rounded-xl pl-12 pr-4 py-3 text-sm text-zinc-300 placeholder:text-zinc-700 outline-none focus:border-zinc-600 transition-colors"/>
  </div>

  <!-- Tabela -->
  <div class="overflow-x-auto rounded-xl border border-zinc-800/40">
    <table class="w-full text-sm">
      <thead>
        <tr class="border-b border-zinc-800/60 bg-[#0e0e0e]">
          <th class="text-left px-5 py-4 text-[10px] uppercase tracking-widest text-zinc-500 font-medium">#</th>
          <th class="text-left px-5 py-4 text-[10px] uppercase tracking-widest text-zinc-500 font-medium">Data</th>
          <th class="text-left px-5 py-4 text-[10px] uppercase tracking-widest text-zinc-500 font-medium">Nome do Negócio</th>
          <th class="text-left px-5 py-4 text-[10px] uppercase tracking-widest text-zinc-500 font-medium">Setor</th>
          <th class="text-left px-5 py-4 text-[10px] uppercase tracking-widest text-zinc-500 font-medium">Faturamento</th>
          <th class="text-left px-5 py-4 text-[10px] uppercase tracking-widest text-zinc-500 font-medium">Ações</th>
        </tr>
      </thead>
      <tbody id="tabela-corpo">
        {% for r in respostas %}
        <tr class="linha-resposta border-b border-zinc-800/20 hover:bg-zinc-900/30 transition-colors"
            id="linha-{{ r['id'] }}"
            data-nome="{{ r['nome_negocio'] }}"
            data-setor="{{ r['setor'] }}"
            data-json="{{ r | tojson | e }}">
          <td class="px-5 py-4 text-zinc-600">{{ r['id'] }}</td>
          <td class="px-5 py-4 text-zinc-500 text-xs">{{ r['data_envio'] }}</td>
          <td class="px-5 py-4 text-zinc-200 font-medium">{{ r['nome_negocio'] or '—' }}</td>
          <td class="px-5 py-4 text-zinc-400">{{ r['setor'] or '—' }}</td>
          <td class="px-5 py-4"><span class="text-[10px] bg-zinc-800 text-zinc-300 px-2 py-1 rounded tracking-wide">{{ r['faturamento'] or '—' }}</span></td>
          <td class="px-5 py-4">
            <div class="flex items-center gap-3">
              <button onclick="verDetalheJSON(this)" class="text-zinc-500 hover:text-zinc-200 transition-colors" title="Ver detalhes">
                <span class="material-symbols-outlined text-base">open_in_new</span>
              </button>
              <form method="POST" action="/admin/apagar/{{ r['id'] }}" onsubmit="return confirm('Apagar esta resposta?')">
                <button type="submit" class="text-zinc-700 hover:text-red-500 transition-colors" title="Apagar">
                  <span class="material-symbols-outlined text-base">delete</span>
                </button>
              </form>
            </div>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>

    {% if not respostas %}
    <div id="estado-vazio" class="text-center py-24 space-y-4">
      <span class="material-symbols-outlined text-5xl text-zinc-800">inbox</span>
      <p class="text-zinc-600 text-sm">Ainda não existem respostas.</p>
    </div>
    {% endif %}
  </div>

</main>

<script>
  // ---- Etiquetas legíveis ----
  const etiquetas = {
    id:"ID", data_envio:"Data de Envio", nome_negocio:"Nome do Negócio",
    setor:"Setor", tempo_mercado:"Tempo de Mercado", faturamento:"Faturamento",
    digital:"Presença Digital", estrutura:"Estrutura Organizacional",
    fluxo:"Fluxo de Clientes", objetivos:"Objetivos",
    visao:"Visão Estratégica", investimento:"Investimento em Marketing",
    ambicao:"Ambição Profissional"
  };

  // ---- Pesquisa ----
  function filtrar() {
    const termo = document.getElementById("pesquisa").value.toLowerCase();
    document.querySelectorAll(".linha-resposta").forEach(linha => {
      const n = linha.dataset.nome.toLowerCase();
      const s = linha.dataset.setor.toLowerCase();
      linha.style.display = (n.includes(termo) || s.includes(termo)) ? "" : "none";
    });
  }

  // ---- Modal ----
  function verDetalheJSON(btn) {
    const linha = btn.closest("tr");
    const dados = JSON.parse(linha.dataset.json);
    const conteudo = document.getElementById("modal-conteudo");
    conteudo.innerHTML = "";
    Object.entries(dados).forEach(([k, v]) => {
      if (!v) return;
      const div = document.createElement("div");
      div.className = "space-y-1 border-b border-zinc-800/40 pb-4";
      div.innerHTML = `<p class="text-[10px] uppercase tracking-widest text-zinc-500">${etiquetas[k]||k}</p><p class="text-zinc-200 text-sm leading-relaxed">${v}</p>`;
      conteudo.appendChild(div);
    });
    document.getElementById("modal").classList.add("aberto");
  }
  function fecharModal() { document.getElementById("modal").classList.remove("aberto"); }
  document.getElementById("modal").addEventListener("click", e => { if(e.target===e.currentTarget) fecharModal(); });

  // ---- Toast ----
  function mostrarToast(msg) {
    const t = document.getElementById("toast");
    document.getElementById("toast-msg").textContent = msg;
    t.classList.add("visivel");
    setTimeout(() => t.classList.remove("visivel"), 4000);
  }

  // ---- Cria linha HTML para nova resposta ----
  function criarLinha(r) {
    const tr = document.createElement("tr");
    tr.className = "linha-resposta linha-nova border-b border-zinc-800/20 hover:bg-zinc-900/30 transition-colors";
    tr.id = "linha-" + r.id;
    tr.dataset.nome  = r.nome_negocio || "";
    tr.dataset.setor = r.setor || "";
    tr.dataset.json  = JSON.stringify(r);
    tr.innerHTML = `
      <td class="px-5 py-4 text-zinc-600">${r.id} <span class="badge-novo ml-1 text-[9px] bg-green-900/60 text-green-400 px-1.5 py-0.5 rounded tracking-widest uppercase">novo</span></td>
      <td class="px-5 py-4 text-zinc-500 text-xs">${r.data_envio}</td>
      <td class="px-5 py-4 text-zinc-200 font-medium">${r.nome_negocio || '—'}</td>
      <td class="px-5 py-4 text-zinc-400">${r.setor || '—'}</td>
      <td class="px-5 py-4"><span class="text-[10px] bg-zinc-800 text-zinc-300 px-2 py-1 rounded tracking-wide">${r.faturamento || '—'}</span></td>
      <td class="px-5 py-4">
        <div class="flex items-center gap-3">
          <button onclick="verDetalheJSON(this)" class="text-zinc-500 hover:text-zinc-200 transition-colors" title="Ver detalhes">
            <span class="material-symbols-outlined text-base">open_in_new</span>
          </button>
          <form method="POST" action="/admin/apagar/${r.id}" onsubmit="return confirm('Apagar esta resposta?')">
            <button type="submit" class="text-zinc-700 hover:text-red-500 transition-colors">
              <span class="material-symbols-outlined text-base">delete</span>
            </button>
          </form>
        </div>
      </td>`;
    return tr;
  }

  // ======================================================
  // ⚡ TEMPO REAL — Socket.IO
  // ======================================================
  const socket = io();

  // Quando recebe nova resposta do servidor
  socket.on("nova_resposta", function(dados) {
    const tbody = document.getElementById("tabela-corpo");
    const vazio = document.getElementById("estado-vazio");

    // Remove o estado "vazio" se existia
    if (vazio) vazio.remove();

    // Insere a nova linha no topo da tabela
    const novaLinha = criarLinha(dados.resposta);
    tbody.insertBefore(novaLinha, tbody.firstChild);

    // Atualiza o contador
    document.getElementById("contador-total").textContent = dados.total;
    document.getElementById("ultima-data").textContent = dados.resposta.data_envio;

    // Mostra notificação
    const nome = dados.resposta.nome_negocio || "Novo cliente";
    mostrarToast("Nova resposta de: " + nome);
  });

  // Quando uma resposta é apagada
  socket.on("resposta_apagada", function(dados) {
    const linha = document.getElementById("linha-" + dados.id);
    if (linha) {
      linha.style.transition = "opacity .4s";
      linha.style.opacity = "0";
      setTimeout(() => linha.remove(), 400);
    }
    document.getElementById("contador-total").textContent = dados.total;
  });

  // Indicador de ligação
  socket.on("connect",    () => { document.getElementById("dot-ligado").className = "w-2 h-2 rounded-full bg-green-500"; });
  socket.on("disconnect", () => { document.getElementById("dot-ligado").className = "w-2 h-2 rounded-full bg-red-500"; });
</script>
</body>
</html>"""


# ============================================================
# INÍCIO
# ============================================================
if __name__ == "__main__":
    criar_tabela()
    print("=" * 52)
    print("  Servidor Strativa iniciado!")
    print("  Formulario:  http://localhost:5000")
    print("  Admin:       http://localhost:5000/admin")
    print("  Para parar:  CTRL + C")
    print("=" * 52)
    port = int(os.environ.get("PORT", 5000))
socketio.run(app, host="0.0.0.0", port=port)
